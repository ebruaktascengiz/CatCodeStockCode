import pandas as pd
import openpyxl
import os
import copy

# --- AYARLAR: Lütfen bu bölümü kendi dosyanıza göre düzenleyin ---

# Orijinal Excel dosyanızın tam yolu (başına 'r' koymayı unutmayın)
ana_dosya_yolu = r'C:\Users\ebru.aktas\Desktop\Cost Codes - all.xlsx'

# Projelerin ayrıştırılacağı sütunun adı
proje_sutunu = 'Code'

# Orijinal verilerin bulunduğu sayfanın adı
ana_sayfa_adi = 'CCC guncel'  # Kendi sayfa adınızı buraya yazın

# Yeni dosyaların kaydedileceği klasör adı
cikis_klasoru = 'Proje Dosyalari Tam Formatli'


# ----------------------------------------------------------------

def kopyala_ve_kaydet():
    """Ana fonksiyon, işlemleri sırasıyla yürütür."""
    print("İşlem başlıyor: Hücre stillerini (kenarlıklar dahil) koruyan optimize edilmiş yöntem.")

    # 1. Çıktı klasörünü oluştur
    if not os.path.exists(cikis_klasoru):
        os.makedirs(cikis_klasoru)
        print(f"'{cikis_klasoru}' adında bir klasör oluşturuldu.")

    try:
        # 2. ADIM (HIZLI): Ana dosyadan sadece benzersiz proje kodlarını almak için pandas kullan
        print("Benzersiz proje kodları okunuyor...")
        df = pd.read_excel(ana_dosya_yolu, sheet_name=ana_sayfa_adi)
        if proje_sutunu not in df.columns:
            raise KeyError(f"'{proje_sutunu}' adında bir sütun bulunamadı. Sütun adını kontrol edin.")
        benzersiz_projeler = df[proje_sutunu].unique()
        # Proje kodunun bulunduğu sütunun index'ini bul (openpyxl 1'den başlar)
        proje_sutun_indexi = df.columns.get_loc(proje_sutunu)

        print(f"Toplam {len(benzersiz_projeler)} adet proje bulundu.")

        # 3. ADIM (OPTİMİZE): Ana dosyayı format okumak için SADECE BİR KEZ aç
        print("Orijinal dosya formatıyla birlikte hafızaya alınıyor...")
        ana_wb = openpyxl.load_workbook(ana_dosya_yolu)
        ana_ws = ana_wb[ana_sayfa_adi]

        # 4. ADIM (FORMAT KORUYARAK): Her proje için dosyaları oluştur
        for proje_kodu in benzersiz_projeler:
            print(f"  -> '{proje_kodu}' projesi işleniyor...")

            # Yeni bir çalışma kitabı ve sayfası oluştur
            yeni_wb = openpyxl.Workbook()
            yeni_ws = yeni_wb.active
            yeni_ws.title = ana_sayfa_adi

            # Başlık satırını tüm stilleriyle kopyala
            baslik_satiri = ana_ws[1]
            yeni_ws.append([cell.value for cell in baslik_satiri])
            for col_idx, cell in enumerate(baslik_satiri, 1):
                yeni_hucre = yeni_ws.cell(row=1, column=col_idx)
                if cell.has_style:
                    yeni_hucre.font = copy.copy(cell.font)
                    yeni_hucre.border = copy.copy(cell.border)
                    yeni_hucre.fill = copy.copy(cell.fill)
                    yeni_hucre.alignment = copy.copy(cell.alignment)

            # İlgili proje koduna sahip veri satırlarını kopyala
            for row in ana_ws.iter_rows(min_row=2):
                # Satırın proje kodunu kontrol et
                if row[proje_sutun_indexi].value == proje_kodu:
                    # Satırı yeni sayfaya ekle (önce değerler, sonra formatlar)
                    yeni_ws.append([cell.value for cell in row])
                    yeni_satir_no = yeni_ws.max_row
                    for col_idx, cell in enumerate(row, 1):
                        yeni_hucre = yeni_ws.cell(row=yeni_satir_no, column=col_idx)
                        if cell.has_style:
                            yeni_hucre.font = copy.copy(cell.font)
                            yeni_hucre.border = copy.copy(cell.border)
                            yeni_hucre.fill = copy.copy(cell.fill)
                            yeni_hucre.alignment = copy.copy(cell.alignment)

            # Sütun genişliklerini kopyala
            for col_letter, dimension in ana_ws.column_dimensions.items():
                yeni_ws.column_dimensions[col_letter].width = dimension.width

            # Yeni dosyayı kaydet
            temiz_dosya_adi = str(proje_kodu).replace('/', '-').replace('\\', '-')
            yeni_dosya_yolu = os.path.join(cikis_klasoru, f"{temiz_dosya_adi}.xlsx")
            yeni_wb.save(yeni_dosya_yolu)

        print("\nİşlem başarıyla tamamlandı!")
        print(f"Tüm dosyalar '{os.path.abspath(cikis_klasoru)}' klasörüne kaydedildi.")

    except FileNotFoundError:
        print(f"\nHATA: '{ana_dosya_yolu}' adında bir dosya bulunamadı.")
    except KeyError as e:
        print(f"\nHATA: {e}. Sayfa veya sütun adı yanlış olabilir. AYARLAR bölümünü kontrol edin.")
    except Exception as e:
        print(f"\nBeklenmedik bir hata oluştu: {e}")


# Fonksiyonu çalıştır
kopyala_ve_kaydet()